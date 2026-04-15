"""
Generates mock Excel data file for FXF3A_Tool.
Sheet structure: Row 1 = title (ignored), Row 2 = headers, Row 3+ = data.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="0F1C3F")
TITLE_FONT   = Font(bold=True, italic=True, color="555E7A")

def write_sheet(ws, title_text, headers, rows):
    """Write title row, header row, then data rows."""
    # Row 1 — title (ignored by loader)
    ws.append([title_text])
    ws["A1"].font = TITLE_FONT

    # Row 2 — headers
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Row 3+ — data
    for row in rows:
        ws.append(row)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)


# ── FXF3A_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3A_Batch")
headers = [
    "ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM",
    "RELEASE?",
    "DISC1","EFF_DATE1","CAN_DATE1",
    "DISC2","EFF_DATE2","CAN_DATE2",
    "DISC3","EFF_DATE3","CAN_DATE3",
    "CURRENCY","INTER","TYPE_HAUL","MATRIX",
    "GEO_DIR1","GEO_TYPE1","GEO_NAME1",
    "PREPD_IN","PREPD_OUT","COLL_IN","COLL_OUT","3RD_PTY",
    "APPLY_ARB","INC_EXM","FAK",
    "N50","N55","N60","N65","N70","N77","N85","N92",
    "N100","N110","N125","N150","N175","N200","N250","N300","N400","N500",
]
rows = [
    # Row with most boolean flags set to Y for testing
    ["A","FXFM","CC","100001","AUTH","001","100","N",
     "10.00","01/01/25","12/31/25","","","","","","","USD","NA","NA","",
     "NA","NA","",
     "Y","N","N","N","N","Y","N","N",
     "Y","N","Y","N","Y","N","Y","N","Y","N","Y","N","Y","N","Y","N","Y","N"],
    # Row with Release=Y and mixed booleans
    ["A","FXFM","CC","100001","AUTH","001","200","Y",
     "15.00","01/01/25","12/31/25","5.00","06/01/25","12/31/25","","","","USD","","","",
     "","","",
     "N","Y","Y","N","N","N","Y","N",
     "N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N"],
    # CHANGE action, no booleans
    ["C","FXFM","CN","100002","AUTH","002","100","N",
     "12.00","01/01/25","","","","","","","","USD","","","",
     "","","",
     "N","N","N","N","N","N","N","N",
     "N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N"],
    # CDN carrier with Release=Y
    ["A","ARFW","CC","200001","AUTH","003","100","Y",
     "20.00","03/01/25","12/31/25","","","","","","","CDN","","","",
     "","","",
     "Y","Y","N","N","Y","N","N","Y",
     "Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y","Y"],
    # DELETE action
    ["D","FXFM","NC","100003","AUTH","001","100","N",
     "","","","","","","","","","","","","",
     "","","",
     "N","N","N","N","N","N","N","N",
     "N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N","N"],
]
write_sheet(ws, "FXF3A — Customer Discount Items", headers, rows)


# ── FXF3B_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3B_Batch")
headers = [
    "ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART",
    "RELEASE?","PREPD_IN","PREPD_OUT","COLL_IN","COLL_OUT",
]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","001","N","Y","N","N","N"],
    ["A","FXFM","CC","100001","AUTH","001","200","001","Y","N","Y","Y","N"],
    ["C","FXFM","CN","100002","AUTH","002","100","001","N","N","N","Y","Y"],
    ["A","ARFW","CC","200001","AUTH","003","100","002","Y","Y","Y","N","N"],
]
write_sheet(ws, "FXF3B — Discounts by State/Terminal", headers, rows)


# ── FXF3C_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3C_Batch")
headers = [
    "ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART",
    "RELEASE?",
]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","001","N"],
    ["A","FXFM","CC","100001","AUTH","001","200","002","Y"],
    ["C","FXFM","CN","100002","AUTH","002","100","001","N"],
    ["D","FXNL","CC","300001","AUTH","004","100","001","N"],
]
write_sheet(ws, "FXF3C — Customer Geography Discounts", headers, rows)


# ── FXF3D_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3D_Batch")
headers = ["ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART"]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","001"],
    ["C","FXFM","CC","100001","AUTH","001","200","001"],
    ["A","ARFW","CN","100002","AUTH","002","100","001"],
]
write_sheet(ws, "FXF3D — Customer Product Discounts", headers, rows)


# ── FXF3E_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3E_Batch")
headers = [
    "ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART",
    "RATE_MANUALLY",
]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","001","N"],
    ["A","FXFM","CC","100002","AUTH","001","100","002","Y"],
    ["C","FXFM","CN","100003","AUTH","002","100","001","N"],
    ["D","FXFM","CC","100001","AUTH","001","200","001","N"],
]
write_sheet(ws, "FXF3E — Customer Rates", headers, rows)


# ── FXF3F_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3F_Batch")
headers = ["ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART"]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","001"],
    ["A","FXFM","CC","100001","AUTH","001","100","002"],
    ["C","ARFW","CN","200001","AUTH","003","100","001"],
]
write_sheet(ws, "FXF3F — Customer Discounts/Adjustments", headers, rows)


# ── FXF3G_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3G_Batch")
headers = [
    "ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART",
    "RELEASE?",
]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","001","N"],
    ["A","FXFM","CC","100001","AUTH","001","100","002","Y"],
    ["C","FXFM","CN","100002","AUTH","002","100","001","N"],
    ["A","FXNL","NN","400001","AUTH","005","100","001","Y"],
]
write_sheet(ws, "FXF3G — Customer Charges/Allowances", headers, rows)


# ── FXF3J_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3J_Batch")
headers = [
    "ACTION","CARRIER","CUST_TYPE",
    "FROM_NAME","FROM_AUTH","FROM_NBR","FROM_ITEM","FROM_PART",
    "TO_TYPE","TO_NAME","TO_AUTH","TO_NBR","TO_ITEM","TO_PART",
    "TO_CARRIER","TO_RELEASE","COPY_EFF_DATE",
]
rows = [
    ["COPY","FXFM","CC","CUST A","AUTH","001","100","001","CC","CUST B","AUTH","002","100","001","FXFM","N","01/01/25"],
    ["COPY","FXFM","CN","CUST C","AUTH","003","100","001","CN","CUST D","AUTH","004","100","001","FXFM","Y","01/01/25"],
]
write_sheet(ws, "FXF3J — Copy Account", headers, rows)


# ── FXF3K_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3K_Batch")
headers = ["ACTION","CARRIER","MATRIX_NAME","MATRIX_EFF_DATE","MATRIX_CANCEL_DATE"]
rows = [
    ["A","FXFM","MATRIX01","01/01/25","12/31/25"],
    ["A","FXFM","MATRIX02","03/01/25","12/31/25"],
    ["C","ARFW","MATRIX03","01/01/25",""],
]
write_sheet(ws, "FXF3K — State Matrix", headers, rows)


# ── FXF3M_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3M_Batch")
headers = [
    "ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART",
    "RELEASE?","RATE_MANUAL",
    "EWR_CLS","EWR_LOW_RATE","EWR_HIGH_RATE","EWR_HIGHEST_VOL_BY_WGT",
]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","001","N","N","N","N","N","N"],
    ["A","FXFM","CC","100001","AUTH","001","100","002","Y","Y","Y","N","N","N"],
    ["C","FXFM","CN","100002","AUTH","002","100","001","N","N","Y","Y","Y","Y"],
]
write_sheet(ws, "FXF3M — Handling Unit Allowance", headers, rows)


# ── FXF3N_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF3N_Batch")
headers = ["ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART"]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","001"],
    ["C","FXFM","CC","100001","AUTH","001","200","001"],
    ["A","ARFW","CN","200001","AUTH","003","100","001"],
    ["D","FXFM","NC","100003","AUTH","001","100","001"],
]
write_sheet(ws, "FXF3N — Unit Rates", headers, rows)


# ── FXF4M_Batch ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("FXF4M_Batch")
headers = [
    "ACTION","CARRIER","CUST_TYPE","ACCOUNT","AUTHORITY","NUMBER","ITEM","PART",
    "EFF_DATE","EXP_DATE",
    "PREPAID_IN","PREPAID_OUT","COLLECT_IN","COLLECT_OUT","THIRD_PARTY",
]
rows = [
    ["A","FXFM","CC","100001","AUTH","001","100","PAY1","01/01/25","12/31/25","Y","N","N","N","N"],
    ["A","FXFM","CC","100001","AUTH","001","200","PAY1","01/01/25","12/31/25","N","Y","Y","N","N"],
    ["C","FXFM","CN","100002","AUTH","002","100","PAY2","01/01/25","","N","N","N","Y","Y"],
    ["A","ARFW","CC","200001","AUTH","003","100","PAY1","03/01/25","12/31/25","Y","Y","Y","Y","Y"],
]
write_sheet(ws, "FXF4M — Earned Discount", headers, rows)


# ── Save ─────────────────────────────────────────────────────────────────────
out_path = r"C:\Users\venkat.govindan\OneDrive - Accenture\personal\Projects\FX3A_Tool\mock_batch_data.xlsx"
wb.save(out_path)
print("Saved:", out_path)
print("Sheets:", wb.sheetnames)
